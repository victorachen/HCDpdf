#start doing electrical load
#map all the park ID's, enter into the HCD 415 app field
#how many trailers can you have in your name? if so, put name in LLC name [ ask HCD on monday -- calendar]

from PyPDF2 import PdfFileReader, PdfFileWriter
import openpyxl
from openpyxl import load_workbook

#takes inputs from excel input file, organizes it into dictionary d
def openpyxl():
    path = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input.xlsx'
    file = load_workbook(path)
    sheet = file.active

    d = {}
    for row in range(sheet.max_row):
        num = sheet.cell(row=row+1, column=2).value
        field = sheet.cell(row=row+1, column=3).value

        if type(num)==int  and field !=None:
            combined = str(num) + "_"+ str(field)
            raw_input = sheet.cell(row=row + 1, column=4).value
            d[combined] = raw_input

    #append a bunch of (hard coded) stuff to the back of the dictionary
    d = append_dic(d)

    print(d)
    return d

#append stuff to the excel-made-dictionary (above)
#ex)from "9_date", we can deduce that the month is "April" and day is "15"; add this stuff to the back of the dictionary
def append_dic(d):
    todays_month = d['9_date'][0:4]
    todays_year = d['9_date'][-5:]
    todays_day = d['9_date'][4:6]
    yr_of_manufacture = d['2_Datefirstsold'][-5:]
    d['todays_month'] = todays_month
    d['todays_year'] = todays_year
    d['todays_day'] = todays_day
    d['yr_of_manufacture'] = yr_of_manufacture
    d['8_llc'] = d['8_llc']+', carrier-- Victor Chen'
    d['6_SitusAddress'] = d['6_SitusAddress'] + ', Yucaipa, CA 92399'
    d['7_Parkname'] = d['7_Parkname'] + ' Mobile Home Park'
    return d

# helper function, that alters PDF (maps whatever is in dictionary)
def alterpdf(emptypath,filledpath):
    d = openpyxl()
    reader = PdfFileReader(emptypath)
    writer = PdfFileWriter()
    totalpages = reader.numPages
    for i in range(totalpages):
        page = reader.pages[i]
        fields = reader.getFields()
        writer.addPage(page)
        for x in d:
            writer.updatePageFormFieldValues(
                writer.getPage(i), {x: d[x]}
            )
        # sorts altered PDF into the "output" folder
        with open(filledpath, "wb") as output_stream:
            writer.write(output_stream)

#for the title trasnfer process
def dupcerttitle():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\dupcerttitle_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\dupcertitle_filled.pdf'
    alterpdf(emptypath,filledpath)
def dupregcard():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\dupregcard_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\dupregcard_filled.pdf'
    alterpdf(emptypath,filledpath)
def billofsale():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\billofsale_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\billofsale_filled.pdf'
    alterpdf(emptypath, filledpath)
def multipurposetransfer():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\multipurposetransfer_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\multipurposetransfer_filled.pdf'
    alterpdf(emptypath, filledpath)

#for the construction permit
def hcd415():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\hcd415_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\hcd415_filled.pdf'
    alterpdf(emptypath, filledpath)
def electricalload():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\electricalload_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\electricalload_filled.pdf'
    alterpdf(emptypath, filledpath)

#extra stuff you might need
def retailvalue():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\retailvalue_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\retailvalue_filled.pdf'
    alterpdf(emptypath, filledpath)
def statementfacts():
    emptypath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\input\statementfacts_empty.pdf'
    filledpath = r'C:\Users\Lenovo\PycharmProjects\HCDpdf\output\statementfacts_filled.pdf'
    alterpdf(emptypath, filledpath)


#combine every file in the filled path
def combine():
    dupcerttitle()
    dupregcard()
    billofsale()
    multipurposetransfer()
    hcd415()
    # electricalload()
    # retailvalue()
    # statementfacts()
    return None

combine()